"""
:author: Thomas Krempf-Driesbach
:date: 2025-11-12
:description:

    Ce scrypte contient la classe permetant de créer et écrire le fichier xlsx

    Il est constituer d'une class outil Xlsx_file contenant les methode de base pour ecrire un fichier xlsx 
    et une classe ProjetXlsx specifique au projet engrenage

    La classe ProjetXlsx utilise les description des objet de type Global pour pouvoir structurer l'ecriture des cellules
"""

import copy
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from modeles2 import Global, Train_simple, Train_epi

NAME = 'reducteur'
PATH = '.\\' # path dans le quelle on souhaite ecrire le fichier
LIGNE_TITRE = 2 # ligne dans la quelle ce trouve le titre principale, est utiliser comme reference pour toutes les autres valeurs
COLONNE_DEPART = 2 # numero de colonne de la premiere valeur
NBR_COLONNE_SEPRARATION = 1 # nombre de colonne separant les train
NBR_COLONNE_TRAIN = 3 # nombre de colonne pris par un train



class XlsxFile():
    def __init__(self,path) -> None:
        """
        contient les methode permetant d'ecrire dans un fichir xlsx

        :param path: chemain complet du fichier xlsx à creer
        """
        self._fichier_excel = path


    def creation_espace_travail(self) -> None:
        """
        creation d'un nouveau fichier excel
        """
        self._wb = Workbook()
        self._ws = self._wb.active


    def ouverture_espace_existant(self) -> None:
        """
        ouvre un fichier excel existant pour l'editer
        """
        self._wb = load_workbook(self._fichier_excel)
        self._ws = self._wb.active


    def trouver_ligne_de_valeur(self,valeur:str,colonne:int,ligne_depart:int) -> int:
        """
        cherche la ligne d'une cellule contenant une valeur donnée

        :param valeur: valeur à chercher
        :param colonne: colonne dans laquelle on cherche
        :param ligne_depart: ligne à partir de laquelle on commence à chercher

        :return: numero de la ligne contenant la valeur, -1 si non trouvé
        """
        for i in range(ligne_depart, self._ws.max_row + 1):
            if self._lire_cellule(i, colonne) == valeur:
                return i
        return -1


    def _ecrire_liste_colonne(self, liste:list[str], ligne_depart:int, colonne:int) -> None:
        """
        ecrire une liste de celle de ligne en ligne en gardant la meme colonne de reference

        :param liste: liste contenant les valeur à écrire, la longueur definit le nbr de cellule a ecrire
        :param ligne_depart: ligne à partir de la quele on commence à ecrire
        :param colonne: colonne de reference dans la quelle toutes les valeurs seront ecrite
        
        **Préconditions :**
        - ``self._ws`` doit etre valide
        """
        for i, valeur in enumerate(liste):
            self._ws.cell(row=ligne_depart + i, column=colonne, value=valeur)


    def _fusionner_cells(self, ligne:int, colonne_depart:int, colonne_fin:int) -> None:
        """
        fusionne des colonnes d'une meme ligne ensemble

        :param ligne: numero de la ligne à fusionner
        :param colone_depart: numero de la colonne determinant le depart de la fusion
        :param colonne_fin: numero de colonne dertminant la fin de la fusion
        
        **Préconditions :**
        - ``self._ws`` doit etre valide
        """
        start_letter = get_column_letter(colonne_depart)
        end_letter = get_column_letter(colonne_fin)
        self._ws.merge_cells(f"{start_letter}{ligne}:{end_letter}{ligne}")


    def save(self) -> None:
        """
        enregistre les modification faite dans le fichier
        
        **Préconditions :**
        - ``self._fichier_excel`` doit contenir le chemain du fichier avec son nom et type
        """
        self._wb.save(self._fichier_excel)


    def _lire_cellule(self,ligne:int,colonne:int) -> str:
        """
        lit la valeur d'une cellule

        :param ligne: numero de la ligne de la cellule à lire
        :param colonne: numero de la colonne de la cellule à lire

        :return: valeur contenue dans la cellule
        """
        return self._ws.cell(row=ligne, column=colonne).value


class XlsxReducteur(XlsxFile):
    def __init__(self,path) -> None:
        """
        créer un fichier xlsx en y écrivant les parametre global

        :param path: chemain complet du fichier xlsx à creer
        """
        super().__init__(path)
        self._param = [0]


    def ecrire_description_simple(self,param:Global,num:int) -> None:
        """
        s'aucupe décrire un parametre avec un description contenant uniquement des valeurs à unitée
        Dans le cas ou il n'est pas encore incorporer dans self._param, il est ecris dans le xlsx
        Dans le cas ou il y est déja, chaque valeur est vérifier, si elle à changer, la cellule est réecrite

        :param param: contient un objet avec une description contenant  uniquement des valeurs à unitée
        :param num: contient le num unique de reference de l'objet qui permet de le situer dans le xlsx
        """
        # definit variable temporaire
        colonne_unitee = COLONNE_DEPART + (num-1)*(NBR_COLONNE_SEPRARATION + NBR_COLONNE_TRAIN)
        colonne_valeur = colonne_unitee + 1
        if num == len(self._param): 
            # cas ou param n'est pas encore incorporee dans self._param
            self._param.append(copy.deepcopy(param))
            self._ecrire_valeur(self._param[num],colonne_unitee)
        elif num < len(self._param):
            # cas ou param est incorporee dans self._param
            for i, key in enumerate(param.description, start=1):
                if param.description[key] != self._param[num].description[key]:
                    ligne = LIGNE_TITRE + i
                    self._ws.cell(row=ligne, column=colonne_valeur, value=param.description[key])
                    self._param[num].description[key] = param.description[key]

    
    def ecrire_description_ogjet_multiple(self,param:Global,num:int) -> None:  
        """
        s'aucupe décrire un parametre a ogjet multiple
        dans param.description on retrouve plusieur objet de type Global
        Dans le cas ou il n'est pas encore incorporer dans self._param, tout les sous objet son ecris sur la meme ligne
        Dans le cas ou il y est déja, chaque valeur est vérifier, si elle à changer, la cellule est réecrite

        :param param: contient un objet avec une description contenant d'autre objet de type Global
        :param num: contient le num unique de reference de l'objet qui permet de le situer dans le xlsx
        """
        # definit variable temporaire
        colonne_unitee = COLONNE_DEPART + (num-1)*(NBR_COLONNE_SEPRARATION + NBR_COLONNE_TRAIN)
        colonne_valeur = colonne_unitee + 1                   
        decalage = 0    
        if num == len(self._param): 
            # cas ou param n'est pas encore incorporee dans self._param
            self._param.append(copy.deepcopy(param.description))
            # ecris chaque sous objet
            for key in self._param[num]:
                self._ecrire_valeur(self._param[num][key],colonne_unitee,decalage)
                decalage = decalage + len(self._param[num][key].unitee) + 1
            # ecris le titre principal
            self._ws.cell(row=LIGNE_TITRE, column=colonne_unitee, value=param.titre)
        elif num < len(self._param):
            # cas ou param est incorporee dans self._param
            param_des = param.description
            for global_key in self._param[num]:
                for i, key in enumerate(param_des[global_key].description, start=1):
                    if param_des[global_key].description[key] != self._param[num][global_key].description[key]:
                        ligne = LIGNE_TITRE + decalage +i
                        self._ws.cell(row=ligne, column=colonne_valeur, value=param_des[global_key].description[key])
                        self._param[num][global_key].description[key] = param_des[global_key].description[key]
                decalage = decalage + len(self._param[num][global_key].unitee) + 1


    def _ecrire_valeur(self,param:Global,colonne_description:int,decalage_ligne:int=0) -> None:
        """
        écris la descritpion d'un objet de type Global dans un fichier xlsx
        ecris 3 colone, nom valeur et unitée
        en ajoutant aussi un titre dans un cellule avec des 3 colone fuisonnee

        :param param: ogjet contenant la descritpion et le titre à ecrire
        :param colonne_description: donne le numero de la colone dans la quelle est écris le nom des valeurs
        :param decalage_ligne: donne le décalage par rapport à chaque sous objet pour les ecrires dans les memes colonnes
        """
        # genere variable temporaire
        ligne = LIGNE_TITRE + decalage_ligne # ligne reference contenant le nom des valeurs
        liste_valeur = list(param.description.values())
        liste_description = list(param.description.keys())
        liste_globale = [param.titre] + liste_description
        # ecris fichier xlsx
        self._ecrire_liste_colonne(liste_globale,ligne,colonne_description) # ecrit titre + description
        self._fusionner_cells(ligne,colonne_description,colonne_description+2)
        self._ecrire_liste_colonne(liste_valeur,ligne+1,colonne_description+1)
        self._ecrire_liste_colonne(param.unitee,ligne+1,colonne_description+2)


    def lire_fichier(self) -> None:
        '''
        lie un fichier xlsx contenant un projet reducteur

        :return: liste contenant les different train du reducteur [0 = placeholder, 1 = train 1, 2 = train 2, ...]
        retourne [0] si le fichier est vide ou pas de la bonne structure
        '''
        LIGNE_TITRE = 2 # ligne dans la quelle ce trouve le titre principale, est utiliser comme reference pour toutes les autres valeurs
        COLONNE_DEPART = 2 # numero de colonne de la premiere valeur
        NBR_COLONNE_SEPRARATION = 1 # nombre de colonne separant les train
        NBR_COLONNE_TRAIN = 3 # nombre de colonne pris par un train
        self._param = []
        colonne = COLONNE_DEPART
        while True:
            # lire le titre principale pour savoir quel type de train il s'agit
            ligne = LIGNE_TITRE
            valeur_cellule = self._lire_cellule(ligne,colonne)
            if valeur_cellule is None:
                return self._param # cas ou la cellule est vide, fin de la lecture
            elif valeur_cellule.startswith('train_simple_'):
                num_part = valeur_cellule[len('train_simple_'):]
                train = Train_simple( int(num_part) )
            elif valeur_cellule.startswith('train_epi_'):
                num_part = valeur_cellule[len('train_epi_'):]
                train = Train_epi( int(num_part) )
            else:
                return self._param # cas ou le titre n'est pas reconnu, fin de la lecture
            train.titre = valeur_cellule
            # lire chaque valeur de la description du train
            for global_key in train.description:
                for key in train.description[global_key].description:
                    ligne = self.trouver_ligne_de_valeur(key,colonne,ligne+1)
                    if ligne != -1:
                        valeur = self._lire_cellule(ligne,colonne+1)
                        train.description[global_key].description[key] = valeur
                ligne = ligne + 1
            self._param.append( train )
            colonne = colonne + NBR_COLONNE_SEPRARATION + NBR_COLONNE_TRAIN
        

if __name__ == '__main__':

    param_global = Global()
    param_global.description['vitesse_entree'] = 3
    test = XlsxReducteur(param_global)
    test.creation_espace_travail()
    train_1 = Train(1)
    test.ecrire_description_simple(train_1,1)
    train_1.description['vitesse_entree'] = 4
    test.ecrire_description_simple(train_1,1)
    train_2 = Train(2)
    train_2.description['rendement'] = 8
    test.ecrire_description_simple(train_2,2)
    test.save()
